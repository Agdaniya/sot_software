from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from services.firebase_client import FirebaseClient
from utils.logger import logger
import os


# ── Output directory ──────────────────────────────────────────────────────────

def get_reports_directory():
    try:
        documents = os.path.join(os.path.expanduser("~"), "Documents")
        d = os.path.join(documents, "SOT_Reports")
        os.makedirs(d, exist_ok=True)
        return d
    except Exception:
        d = os.path.join(os.path.expanduser("~"), "SOT_Reports")
        os.makedirs(d, exist_ok=True)
        return d


# ── Design tokens (shared with employee report) ───────────────────────────────

def F(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

NAVY       = F("1e2a3a")
NAVY_LIGHT = F("253447")
SURFACE    = F("FFFFFF")
BG         = F("F7F8FA")
BG_ALT     = F("EEF1F5")
MUTED_ROW  = F("F0F4F8")

S_FILL = {
    "not_started":     F("F3F4F6"),
    "in_progress":     F("FEF9C3"),
    "submitted":       F("EDE9FE"),
    "admin_rejected":  F("FFE4E6"),
    "admin_approved":  F("CCFBF1"),
    "client_approved": F("CCFBF1"),
    "completed":       F("CCFBF1"),
}
S_FONT = {
    "not_started":     "6B7280",
    "in_progress":     "92400E",
    "submitted":       "5B21B6",
    "admin_rejected":  "BE123C",
    "admin_approved":  "0F766E",
    "client_approved": "0F766E",
    "completed":       "0F766E",
}
S_LABEL = {
    "not_started":     "Not Started",
    "in_progress":     "In Progress",
    "submitted":       "Submitted",
    "admin_rejected":  "Rejected",
    "admin_approved":  "Admin Approved",
    "client_approved": "Client Approved",
    "completed":       "Completed",
}

NO_BORDER  = Border()
THIN       = Side(style="thin", color="E5E7EB")
ROW_BORDER = Border(bottom=THIN)
NCOLS      = 9   # A–I


# ── Shared helpers ────────────────────────────────────────────────────────────

def _fill_row(ws, row, bg, border=ROW_BORDER):
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = bg
        c.border = border


def _cell(ws, row, col, value="", fg="1F2328", size=10, bold=False,
          italic=False, bg=SURFACE, border=ROW_BORDER,
          ah="left", av="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, italic=italic, color=fg, size=size)
    c.alignment = Alignment(horizontal=ah, vertical=av, wrap_text=wrap)
    c.border    = border
    c.fill      = bg
    return c


def _merge(ws, row, c1, c2):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def _cover_block(ws, row, title, subtitle, generated_on):
    _fill_row(ws, row,     NAVY, NO_BORDER); ws.row_dimensions[row].height     = 10
    _fill_row(ws, row + 1, NAVY, NO_BORDER); ws.row_dimensions[row + 1].height = 32
    _fill_row(ws, row + 2, NAVY, NO_BORDER); ws.row_dimensions[row + 2].height = 22

    c = _cell(ws, row + 1, 2, title, fg="FFFFFF", size=16, bold=True, bg=NAVY, border=NO_BORDER)
    _merge(ws, row + 1, 2, NCOLS)
    _cell(ws, row + 2, 2, subtitle, fg="94A3B8", size=10, bg=NAVY, border=NO_BORDER)
    _merge(ws, row + 2, 2, NCOLS)

    _fill_row(ws, row + 3, F("3B82F6"), NO_BORDER); ws.row_dimensions[row + 3].height = 3

    _fill_row(ws, row + 4, MUTED_ROW, NO_BORDER)
    _cell(ws, row + 4, 2, f"Generated: {generated_on}",
          fg="64748B", size=9, bg=MUTED_ROW, border=NO_BORDER)
    _merge(ws, row + 4, 2, NCOLS)
    ws.row_dimensions[row + 4].height = 18

    _fill_row(ws, row + 5, SURFACE, NO_BORDER); ws.row_dimensions[row + 5].height = 8
    return row + 6


def _project_banner(ws, row, name, client):
    _fill_row(ws, row, NAVY_LIGHT, NO_BORDER)
    _cell(ws, row, 2, name.upper(), fg="FFFFFF", size=11, bold=True,
          bg=NAVY_LIGHT, border=NO_BORDER)
    _merge(ws, row, 2, 5)
    _cell(ws, row, 6, client, fg="94A3B8", size=10,
          bg=NAVY_LIGHT, border=NO_BORDER, ah="right")
    _merge(ws, row, 6, NCOLS)
    ws.row_dimensions[row].height = 24
    return row + 1


def _meta_row(ws, row, project_id, assigned_text):
    _fill_row(ws, row, MUTED_ROW, NO_BORDER)
    _cell(ws, row, 2, f"Project ID: {project_id}",
          fg="64748B", size=8, italic=True, bg=MUTED_ROW, border=NO_BORDER)
    _merge(ws, row, 2, 4)
    _cell(ws, row, 5, f"Assigned: {assigned_text}",
          fg="64748B", size=8, italic=True, bg=MUTED_ROW, border=NO_BORDER, ah="right")
    _merge(ws, row, 5, NCOLS)
    ws.row_dimensions[row].height = 16
    return row + 1


def _section_label(ws, row, text):
    _fill_row(ws, row, SURFACE, NO_BORDER)
    _cell(ws, row, 2, text.upper(), fg="64748B", size=8, bold=True, bg=SURFACE,
          border=Border(bottom=Side(style="thin", color="3B82F6")))
    _merge(ws, row, 2, NCOLS)
    ws.row_dimensions[row].height = 20
    return row + 1


def _col_headers(ws, row, headers: dict):
    _fill_row(ws, row, BG, ROW_BORDER)
    for col, label in headers.items():
        _cell(ws, row, col, label, fg="6B7280", size=9, bold=True,
              bg=BG, border=ROW_BORDER, ah="center")
    ws.row_dimensions[row].height = 18
    return row + 1


def _status_cell(ws, row, col, status_key):
    label = S_LABEL.get(status_key, status_key.replace("_", " ").title())
    c = ws.cell(row=row, column=col, value=label)
    c.fill      = S_FILL.get(status_key, BG)
    c.font      = Font(name="Calibri", bold=True, size=9,
                       color=S_FONT.get(status_key, "6B7280"))
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = ROW_BORDER


def _pay_cell(ws, row, col, pay_txt, paid, bg_row):
    c = ws.cell(row=row, column=col, value=pay_txt)
    if pay_txt == "—":
        c.fill = bg_row; c.font = Font(name="Calibri", color="9CA3AF", size=9)
    elif paid:
        c.fill = F("CCFBF1"); c.font = Font(name="Calibri", bold=True, color="0F766E", size=9)
    else:
        c.fill = F("FFE4E6"); c.font = Font(name="Calibri", bold=True, color="BE123C", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = ROW_BORDER


def _summary_bar(ws, row, parts: list):
    _fill_row(ws, row, MUTED_ROW, NO_BORDER)
    text = "    ·    ".join(parts)
    _cell(ws, row, 2, f"  {text}", fg="374151", size=9, bold=True,
          bg=MUTED_ROW, border=NO_BORDER)
    _merge(ws, row, 2, NCOLS)
    ws.row_dimensions[row].height = 18
    return row + 1


def _gap(ws, row, height=10):
    _fill_row(ws, row, SURFACE, NO_BORDER)
    ws.row_dimensions[row].height = height
    return row + 1


def _divider(ws, row):
    _fill_row(ws, row, F("E5E7EB"), NO_BORDER)
    ws.row_dimensions[row].height = 2
    return row + 1


def _fmt_date(d):
    if not d: return "—"
    try:
        return datetime.fromisoformat(d).strftime("%d %b %Y  %H:%M")
    except Exception:
        return str(d)


# ── Main report ───────────────────────────────────────────────────────────────

def generate_project_status_report():
    fb       = FirebaseClient()
    projects = fb.root.child("projects").get() or {}
    drawings = fb.root.child("drawings").get() or {}
    users    = fb.root.child("users").get()    or {}

    wb       = Workbook()
    now_str  = datetime.now().strftime("%B %d, %Y  %H:%M")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    for letter, width in {
        "A": 2, "B": 28, "C": 18, "D": 14,
        "E": 14, "F": 14, "G": 14, "H": 14, "I": 18,
    }.items():
        ws_sum.column_dimensions[letter].width = width

    srow = _cover_block(ws_sum, 1,
                        "Project Status Report",
                        "Staff Operations Tracker  ·  SOT",
                        now_str)

    srow = _col_headers(ws_sum, srow, {
        2: "Project", 3: "Client",
        4: "Total Drawings", 5: "Not Started", 6: "In Progress",
        7: "Submitted", 8: "Client Approved", 9: "Completion %",
    })

    sorted_projects = sorted(projects.items(),
                             key=lambda x: x[1].get("name", "").lower())

    for idx, (project_id, project) in enumerate(sorted_projects):
        proj_drawings = drawings.get(project_id, {}) or {}
        if not isinstance(proj_drawings, dict):
            proj_drawings = {}

        counts = {k: 0 for k in (
            "not_started", "in_progress", "submitted",
            "admin_approved", "admin_rejected", "client_approved", "completed"
        )}
        for d in proj_drawings.values():
            s = d.get("status", "not_started")
            if s in counts:
                counts[s] += 1

        total      = len(proj_drawings)
        approved   = counts["client_approved"] + counts["admin_approved"] + counts["completed"]
        comp_pct   = f"{round(approved / total * 100, 1)}%" if total else "0%"

        bg = BG_ALT if idx % 2 else SURFACE
        _fill_row(ws_sum, srow, bg, ROW_BORDER)
        _cell(ws_sum, srow, 2, project.get("name", ""),
              fg="1F2328", size=10, bold=True, bg=bg)
        _cell(ws_sum, srow, 3, project.get("client_name", ""),
              fg="6B7280", size=9, bg=bg)
        _cell(ws_sum, srow, 4, total or "—",
              fg="374151", size=10, bg=bg, ah="center")
        _cell(ws_sum, srow, 5, counts["not_started"] or "—",
              fg="6B7280", size=10, bg=bg, ah="center")

        # In Progress — highlight if any
        ip = counts["in_progress"]
        ip_bg = F("FEF9C3") if ip else bg
        c = ws_sum.cell(row=srow, column=6, value=ip or "—")
        c.fill = ip_bg; c.border = ROW_BORDER
        c.font = Font(name="Calibri", bold=bool(ip),
                      color="92400E" if ip else "9CA3AF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Submitted — highlight if any
        sb = counts["submitted"]
        sb_bg = F("EDE9FE") if sb else bg
        c = ws_sum.cell(row=srow, column=7, value=sb or "—")
        c.fill = sb_bg; c.border = ROW_BORDER
        c.font = Font(name="Calibri", bold=bool(sb),
                      color="5B21B6" if sb else "9CA3AF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Client approved — teal if any
        ca = counts["client_approved"]
        ca_bg = F("CCFBF1") if ca else bg
        c = ws_sum.cell(row=srow, column=8, value=ca or "—")
        c.fill = ca_bg; c.border = ROW_BORDER
        c.font = Font(name="Calibri", bold=bool(ca),
                      color="0F766E" if ca else "9CA3AF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

        _cell(ws_sum, srow, 9, comp_pct,
              fg="1F2328", size=10, bold=True, bg=bg, ah="center")

        ws_sum.row_dimensions[srow].height = 20
        srow += 1

    # ── Sheet 2: Detail (collapsible per-project) ─────────────────────────────
    ws = wb.create_sheet(title="Detail")
    ws.sheet_view.showGridLines = False
    # summaryBelow=False → the [−] collapse button sits on the banner row itself
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    for letter, width in {
        "A": 2, "B": 28, "C": 16, "D": 14,
        "E": 14, "F": 16, "G": 20, "H": 20, "I": 16,
    }.items():
        ws.column_dimensions[letter].width = width

    row = _cover_block(ws, 1,
                       "Project Status — Detail",
                       "Drawing-level breakdown for all projects  ·  click [+] on the left to expand a project",
                       now_str)

    STATUS_ORDER = {
        "in_progress": 0, "submitted": 1, "not_started": 2,
        "admin_rejected": 3, "admin_approved": 4,
        "client_approved": 5, "completed": 6,
    }

    for project_id, project in sorted_projects:
        proj_drawings = drawings.get(project_id, {}) or {}
        if not isinstance(proj_drawings, dict):
            proj_drawings = {}

        # Assigned employees
        assigned = project.get("assigned_users", {})
        names = [users.get(uid, {}).get("username", uid)
                 for uid in (assigned.keys() if isinstance(assigned, dict) else [])]
        assigned_text = ", ".join(names) if names else "None"

        # Banner row is level 0 — always visible, holds the [+/−] button
        banner_row = row
        row = _project_banner(ws, row, project.get("name", ""), project.get("client_name", ""))
        ws.row_dimensions[banner_row].outline_level = 0

        # Everything inside the project (meta, headers, drawings, summary, gap)
        # gets outline_level=1 so it collapses under the banner
        counts = {k: 0 for k in (
            "not_started", "in_progress", "submitted",
            "admin_approved", "admin_rejected", "client_approved", "completed"
        )}
        paid_count = unpaid_count = 0

        meta_row = row
        row = _meta_row(ws, row, project_id, assigned_text)
        ws.row_dimensions[meta_row].outline_level = 1

        sec_row = row
        row = _section_label(ws, row, "Drawings")
        ws.row_dimensions[sec_row].outline_level = 1

        hdr_row = row
        row = _col_headers(ws, row, {
            2: "Drawing Name", 3: "Status", 4: "Progress",
            5: "Steps Done", 6: "Payment", 7: "Payment Date", 8: "Approved Date",
        })
        ws.row_dimensions[hdr_row].outline_level = 1

        if not proj_drawings:
            _fill_row(ws, row, SURFACE, ROW_BORDER)
            c = ws.cell(row=row, column=2, value="No drawings created yet")
            c.font = Font(name="Calibri", italic=True, color="9CA3AF", size=9)
            c.fill = SURFACE; c.border = ROW_BORDER
            c.alignment = Alignment(vertical="center")
            _merge(ws, row, 2, 8)
            ws.row_dimensions[row].height = 18
            ws.row_dimensions[row].outline_level = 1
            row += 1
        else:
            sorted_drawings = sorted(
                proj_drawings.items(),
                key=lambda x: STATUS_ORDER.get(x[1].get("status", "not_started"), 9)
            )

            for d_idx, (drawing_id, drawing) in enumerate(sorted_drawings):
                status           = drawing.get("status", "not_started")
                payment_received = drawing.get("payment_received", False)
                payment_date     = drawing.get("payment_date", "")
                approved_date    = drawing.get("client_approved_date", "")

                if status in counts:
                    counts[status] += 1

                sub_steps = drawing.get("sub_steps", {})
                if sub_steps:
                    done  = sum(1 for s in sub_steps.values() if s.get("completed", False))
                    total = len(sub_steps)
                    progress  = f"{round(done / total * 100)}%"
                    steps_txt = f"{done} / {total}"
                else:
                    progress  = "—"
                    steps_txt = "—"

                if status in ("admin_approved", "client_approved", "completed"):
                    if payment_received:
                        pay_txt = "Paid ✓"; paid_count += 1; paid = True
                    else:
                        pay_txt = "Unpaid ✗"; unpaid_count += 1; paid = False
                else:
                    pay_txt = "—"; paid = None

                bg = BG_ALT if d_idx % 2 else SURFACE
                _fill_row(ws, row, bg, ROW_BORDER)
                _cell(ws, row, 2, drawing.get("name", ""),
                      fg="1F2328", size=10, bold=True, bg=bg)
                _status_cell(ws, row, 3, status)
                _cell(ws, row, 4, progress,  bg=bg, ah="center", fg="374151", bold=True)
                _cell(ws, row, 5, steps_txt, bg=bg, ah="center", fg="6B7280")
                _pay_cell(ws, row, 6, pay_txt,
                          paid=paid if paid is not None else False, bg_row=bg)
                _cell(ws, row, 7, _fmt_date(payment_date),
                      bg=bg, ah="center", fg="6B7280", size=9)
                _cell(ws, row, 8, _fmt_date(approved_date),
                      bg=bg, ah="center", fg="6B7280", size=9)
                ws.row_dimensions[row].height = 20
                ws.row_dimensions[row].outline_level = 1
                row += 1

        # Summary bar — outline_level=1 so it also collapses with the project
        total_drw = len(proj_drawings)
        approved  = counts["client_approved"] + counts["admin_approved"] + counts["completed"]
        comp_pct  = f"{round(approved / total_drw * 100, 1)}%" if total_drw else "0%"
        pay_pct   = (f"{round(paid_count / (paid_count + unpaid_count) * 100, 1)}%"
                     if (paid_count + unpaid_count) else "N/A")

        parts = [f"{total_drw} drawing{'s' if total_drw != 1 else ''}"]
        if counts["in_progress"]:    parts.append(f"{counts['in_progress']} in progress")
        if counts["submitted"]:      parts.append(f"{counts['submitted']} submitted")
        if counts["admin_rejected"]: parts.append(f"{counts['admin_rejected']} rejected")
        if approved:                 parts.append(f"{approved} approved")
        parts.append(f"completion {comp_pct}")
        if paid_count + unpaid_count:
            parts.append(f"payment {pay_pct}  ({paid_count} paid / {unpaid_count} unpaid)")

        sum_row = row
        row = _summary_bar(ws, row, parts)
        ws.row_dimensions[sum_row].outline_level = 1

        gap_row = row
        row = _gap(ws, row, 8)
        ws.row_dimensions[gap_row].outline_level = 1

    # ── Save ─────────────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"project_status_{timestamp}.xlsx"
    full_path = os.path.join(get_reports_directory(), filename)
    wb.save(full_path)
    logger.info(f"Generated project status report: {full_path}")

    # Auto-open
    try:
        os.startfile(full_path)
    except Exception:
        pass

    return full_path
