from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from services.firebase_client import FirebaseClient
from utils.logger import logger
import os


# ── Output directory ──────────────────────────────────────────────────────────

def get_reports_directory():
    try:
        documents = os.path.join(os.path.expanduser("~"), "Documents")
        d = os.path.join(documents, "SOT_Reports")
        os.makedirs(d, exist_ok=True)
        return d
    except Exception:
        d = os.path.join(os.path.expanduser("~"), "SOT_Reports")
        os.makedirs(d, exist_ok=True)
        return d


# ── Design tokens ─────────────────────────────────────────────────────────────

def F(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

# Core palette — matches SOT app design system
NAVY        = F("1e2a3a")   # primary accent (topbar / buttons)
NAVY_LIGHT  = F("253447")   # slightly lighter navy for section dividers
SURFACE     = F("FFFFFF")   # white
BG          = F("F7F8FA")   # page background tint
BG_ALT      = F("EEF1F5")   # alternating row tint
BORDER_COL  = F("E5E7EB")   # border colour
MUTED_ROW   = F("F0F4F8")   # meta / summary highlight

# Status colours
S_FILL = {
    "not_started":     F("F3F4F6"),
    "in_progress":     F("FEF9C3"),
    "submitted":       F("EDE9FE"),
    "admin_rejected":  F("FFE4E6"),
    "admin_approved":  F("CCFBF1"),
    "client_approved": F("CCFBF1"),
    "completed":       F("CCFBF1"),
}
S_FONT = {
    "not_started":     "6B7280",
    "in_progress":     "92400E",
    "submitted":       "5B21B6",
    "admin_rejected":  "BE123C",
    "admin_approved":  "0F766E",
    "client_approved": "0F766E",
    "completed":       "0F766E",
}
S_LABEL = {
    "not_started":     "Not Started",
    "in_progress":     "In Progress",
    "submitted":       "Submitted",
    "admin_rejected":  "Rejected",
    "admin_approved":  "Admin Approved",
    "client_approved": "Client Approved",
    "completed":       "Completed",
}

NO_BORDER    = Border()
THIN         = Side(style="thin",   color="E5E7EB")
THICK_BOTTOM = Side(style="medium", color="1e2a3a")
ROW_BORDER   = Border(bottom=THIN)
SECTION_BORDER = Border(bottom=Side(style="thin", color="1e2a3a"))

NCOLS = 9   # A–I


# ── Low-level helpers ─────────────────────────────────────────────────────────

def _fill_row(ws, row, bg, border=ROW_BORDER):
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = bg
        c.border = border


def _cell(ws, row, col, value="", fg="1F2328", size=10, bold=False,
          italic=False, bg=SURFACE, border=ROW_BORDER,
          ah="left", av="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, italic=italic,
                       color=fg, size=size)
    c.alignment = Alignment(horizontal=ah, vertical=av, wrap_text=wrap)
    c.border    = border
    c.fill      = bg
    return c


def _merge(ws, row, c1, c2):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row,   end_column=c2)


# ── Block builders ────────────────────────────────────────────────────────────

def _cover_block(ws, row, title, subtitle, generated_on, date_range_str):
    """Full-width cover header for the sheet."""
    # Tall navy banner
    _fill_row(ws, row, NAVY, NO_BORDER)
    _fill_row(ws, row + 1, NAVY, NO_BORDER)
    _fill_row(ws, row + 2, NAVY, NO_BORDER)
    ws.row_dimensions[row].height     = 10
    ws.row_dimensions[row + 1].height = 32
    ws.row_dimensions[row + 2].height = 22

    c = _cell(ws, row + 1, 2, title, fg="FFFFFF", size=16, bold=True,
              bg=NAVY, border=NO_BORDER)
    _merge(ws, row + 1, 2, NCOLS)

    c2 = _cell(ws, row + 2, 2, subtitle, fg="94A3B8", size=10,
               bg=NAVY, border=NO_BORDER)
    _merge(ws, row + 2, 2, NCOLS)

    # Thin accent line
    _fill_row(ws, row + 3, F("3B82F6"), NO_BORDER)
    ws.row_dimensions[row + 3].height = 3

    # Meta row
    _fill_row(ws, row + 4, MUTED_ROW, NO_BORDER)
    _cell(ws, row + 4, 2, f"Generated: {generated_on}",
          fg="64748B", size=9, bg=MUTED_ROW, border=NO_BORDER)
    _merge(ws, row + 4, 2, 5)
    dr = _cell(ws, row + 4, 6, date_range_str,
               fg="64748B", size=9, bg=MUTED_ROW,
               border=NO_BORDER, ah="right")
    _merge(ws, row + 4, 6, NCOLS)
    ws.row_dimensions[row + 4].height = 18

    _fill_row(ws, row + 5, SURFACE, NO_BORDER)
    ws.row_dimensions[row + 5].height = 8
    return row + 6


def _employee_banner(ws, row, username, email, role):
    """Navy band with employee name + right-aligned email / role."""
    _fill_row(ws, row, NAVY_LIGHT, NO_BORDER)
    _cell(ws, row, 2, username.upper(), fg="FFFFFF", size=11, bold=True,
          bg=NAVY_LIGHT, border=NO_BORDER)
    _merge(ws, row, 2, 5)
    info = f"{email}   ·   {role}"
    _cell(ws, row, 6, info, fg="94A3B8", size=9,
          bg=NAVY_LIGHT, border=NO_BORDER, ah="right")
    _merge(ws, row, 6, NCOLS)
    ws.row_dimensions[row].height = 24
    return row + 1


def _section_label(ws, row, text, col_start=2):
    """Small ALL-CAPS section label with a bottom accent line."""
    _fill_row(ws, row, SURFACE, NO_BORDER)
    _cell(ws, row, col_start, text.upper(),
          fg="64748B", size=8, bold=True, bg=SURFACE,
          border=Border(bottom=Side(style="thin", color="3B82F6")))
    _merge(ws, row, col_start, NCOLS)
    ws.row_dimensions[row].height = 20
    return row + 1


def _col_headers(ws, row, headers: dict):
    """Light gray column header row."""
    _fill_row(ws, row, BG, ROW_BORDER)
    for col, label in headers.items():
        _cell(ws, row, col, label, fg="6B7280", size=9, bold=True,
              bg=BG, border=ROW_BORDER, ah="center")
    ws.row_dimensions[row].height = 18
    return row + 1


def _status_cell(ws, row, col, status_key, bg_row):
    label = S_LABEL.get(status_key, status_key.replace("_", " ").title())
    c = ws.cell(row=row, column=col, value=label)
    c.fill      = S_FILL.get(status_key, BG)
    c.font      = Font(name="Calibri", bold=True, size=9,
                       color=S_FONT.get(status_key, "6B7280"))
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = ROW_BORDER


def _pay_cell(ws, row, col, pay_txt, paid, bg_row):
    c = ws.cell(row=row, column=col, value=pay_txt)
    if pay_txt == "—":
        c.fill = bg_row
        c.font = Font(name="Calibri", color="9CA3AF", size=9)
    elif paid:
        c.fill = F("CCFBF1")
        c.font = Font(name="Calibri", bold=True, color="0F766E", size=9)
    else:
        c.fill = F("FFE4E6")
        c.font = Font(name="Calibri", bold=True, color="BE123C", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = ROW_BORDER


def _summary_bar(ws, row, text):
    _fill_row(ws, row, MUTED_ROW, NO_BORDER)
    _cell(ws, row, 2, text, fg="374151", size=9, bold=True,
          bg=MUTED_ROW, border=NO_BORDER)
    _merge(ws, row, 2, NCOLS)
    ws.row_dimensions[row].height = 18
    return row + 1


def _gap(ws, row, height=10):
    _fill_row(ws, row, SURFACE, NO_BORDER)
    ws.row_dimensions[row].height = height
    return row + 1


def _divider(ws, row):
    _fill_row(ws, row, F("E5E7EB"), NO_BORDER)
    ws.row_dimensions[row].height = 2
    return row + 1


# ── Main report ───────────────────────────────────────────────────────────────

def generate_employee_performance_report(start_date=None, end_date=None):
    fb           = FirebaseClient()
    users        = fb.root.child("users").get()         or {}
    logs         = fb.root.child("activity_logs").get() or {}
    projects_all = fb.root.child("projects").get()      or {}
    drawings_all = fb.root.child("drawings").get()      or {}

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    for letter, width in {
        "A": 2, "B": 26, "C": 14, "D": 14,
        "E": 14, "F": 14, "G": 14, "H": 14, "I": 14,
    }.items():
        ws_sum.column_dimensions[letter].width = width

    now_str    = datetime.now().strftime("%B %d, %Y  %H:%M")
    dr_str     = ""
    if start_date and end_date:
        dr_str = f"Period: {start_date}  →  {end_date}"
    elif start_date:
        dr_str = f"From: {start_date}"
    elif end_date:
        dr_str = f"Until: {end_date}"
    else:
        dr_str = "All time"

    srow = _cover_block(ws_sum, 1,
                        "Employee Performance Report",
                        "Staff Operations Tracker  ·  SOT",
                        now_str, dr_str)

    # Column headers
    srow = _col_headers(ws_sum, srow, {
        2: "Employee", 3: "Role",
        4: "Days Worked", 5: "Total Hours", 6: "Avg Hours/Day",
        7: "Total Drawings", 8: "In Progress", 9: "Submitted",
    })

    sorted_users = sorted(
        [(uid, u) for uid, u in users.items()],
        key=lambda x: x[1].get("username", "").lower()
    )

    for idx, (user_id, user) in enumerate(sorted_users):
        username = user.get("username", "Unknown")
        role     = user.get("role", "").replace("_", " ").title()

        # Attendance totals
        user_logs = logs.get(user_id, {}) or {}
        if not isinstance(user_logs, dict):
            user_logs = {}
        th = tm = days = 0
        for log_date, entry in user_logs.items():
            if start_date and log_date < start_date: continue
            if end_date   and log_date > end_date:   continue
            hw = entry.get("total_hours", "N/A")
            if hw not in ("N/A", "—", None):
                try:
                    h, m = str(hw).split(":")
                    th += int(h); tm += int(m); days += 1
                except Exception:
                    pass
        extra = tm // 60; rem = tm % 60; total_h = th + extra
        hours_str = f"{total_h}h {rem:02d}m" if days else "—"
        avg_str   = f"{round(total_h / days, 1)}h" if days else "—"

        # Drawing counts
        counts = {"in_progress": 0, "submitted": 0, "total": 0}
        for proj in projects_all.values():
            if user_id not in proj.get("assigned_users", {}):
                continue
            for d in drawings_all.get(
                    next((pid for pid, p in projects_all.items() if p is proj), ""), {}
            ).values():
                pass
        # cleaner loop
        for project_id, project in projects_all.items():
            if user_id not in project.get("assigned_users", {}):
                continue
            for drawing in drawings_all.get(project_id, {}).values():
                s = drawing.get("status", "")
                counts["total"] += 1
                if s == "in_progress":  counts["in_progress"] += 1
                if s == "submitted":    counts["submitted"]   += 1

        bg = BG_ALT if idx % 2 else SURFACE
        _fill_row(ws_sum, srow, bg, ROW_BORDER)
        _cell(ws_sum, srow, 2, username,       fg="1F2328", size=10, bold=True, bg=bg)
        _cell(ws_sum, srow, 3, role,           fg="6B7280", size=9,  bg=bg, ah="center")
        _cell(ws_sum, srow, 4, days or "—",    fg="374151", size=10, bg=bg, ah="center")
        _cell(ws_sum, srow, 5, hours_str,      fg="374151", size=10, bg=bg, ah="center")
        _cell(ws_sum, srow, 6, avg_str,        fg="374151", size=10, bg=bg, ah="center")
        _cell(ws_sum, srow, 7, counts["total"] or "—",
              fg="374151", size=10, bg=bg, ah="center")

        c_ip = counts["in_progress"]
        c_sb = counts["submitted"]
        ip_bg = F("FEF9C3") if c_ip else bg
        sb_bg = F("EDE9FE") if c_sb else bg
        c = ws_sum.cell(row=srow, column=8, value=c_ip or "—")
        c.fill = ip_bg
        c.font = Font(name="Calibri", bold=bool(c_ip),
                      color="92400E" if c_ip else "9CA3AF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = ROW_BORDER

        c = ws_sum.cell(row=srow, column=9, value=c_sb or "—")
        c.fill = sb_bg
        c.font = Font(name="Calibri", bold=bool(c_sb),
                      color="5B21B6" if c_sb else "9CA3AF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = ROW_BORDER

        ws_sum.row_dimensions[srow].height = 20
        srow += 1

    # ── Sheet 2: Detail ───────────────────────────────────────────────────────
    ws = wb.create_sheet(title="Detail")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    for letter, width in {
        "A": 2, "B": 24, "C": 20, "D": 14,
        "E": 12, "F": 12, "G": 14, "H": 16, "I": 12,
    }.items():
        ws.column_dimensions[letter].width = width

    row = _cover_block(ws, 1,
                       "Employee Performance — Detail",
                       "Attendance logs and drawing assignments per employee",
                       now_str, dr_str)

    STATUS_ORDER = {
        "in_progress": 0, "admin_rejected": 1, "not_started": 2,
        "submitted": 3, "admin_approved": 4, "client_approved": 5,
    }

    for user_id, user in sorted_users:
        username = user.get("username", "Unknown")
        email    = user.get("email", "")
        role     = user.get("role", "").replace("_", " ").title()

        row = _employee_banner(ws, row, username, email, role)

        # ── Attendance ────────────────────────────────────────────────────
        row = _section_label(ws, row, "Attendance Log")
        row = _col_headers(ws, row, {
            2: "Date", 3: "Login", 4: "Logout",
            5: "Hours Worked", 6: "Status",
        })

        user_logs = logs.get(user_id, {}) or {}
        if not isinstance(user_logs, dict):
            user_logs = {}

        total_hours = total_mins = days_worked = 0
        has_att = False

        for a_idx, log_date in enumerate(sorted(user_logs.keys())):
            if start_date and log_date < start_date: continue
            if end_date   and log_date > end_date:   continue
            entry        = user_logs[log_date]
            login        = entry.get("login_time",  "—")
            logout       = entry.get("logout_time", "—")
            hours_worked = entry.get("total_hours", "—")

            if logout == "—" or logout is None:
                s_txt = "Active";     s_bg = F("FEF9C3"); s_fg = "92400E"
            elif hours_worked not in ("—", "N/A", None):
                s_txt = "Complete";   s_bg = F("CCFBF1"); s_fg = "0F766E"
            else:
                s_txt = "Incomplete"; s_bg = F("FFE4E6"); s_fg = "BE123C"

            if hours_worked not in ("—", "N/A", None):
                try:
                    h, m = str(hours_worked).split(":")
                    total_hours += int(h); total_mins += int(m); days_worked += 1
                except Exception:
                    pass

            bg = BG_ALT if a_idx % 2 else SURFACE
            _fill_row(ws, row, bg, ROW_BORDER)
            _cell(ws, row, 2, log_date,     bg=bg, ah="center", fg="374151")
            _cell(ws, row, 3, login,        bg=bg, ah="center", fg="374151")
            _cell(ws, row, 4, logout,       bg=bg, ah="center", fg="374151")
            _cell(ws, row, 5, hours_worked, bg=bg, ah="center", fg="374151", bold=True)
            sc = ws.cell(row=row, column=6, value=s_txt)
            sc.fill = s_bg; sc.border = ROW_BORDER
            sc.font = Font(name="Calibri", bold=True, color=s_fg, size=9)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 18
            ws.row_dimensions[row].outline_level = 1
            row += 1
            has_att = True

        if not has_att:
            _fill_row(ws, row, SURFACE, ROW_BORDER)
            c = ws.cell(row=row, column=2, value="No attendance records in this date range")
            c.font = Font(name="Calibri", italic=True, color="9CA3AF", size=9)
            c.fill = SURFACE; c.border = ROW_BORDER
            c.alignment = Alignment(vertical="center")
            _merge(ws, row, 2, 7)
            ws.row_dimensions[row].height = 18
            ws.row_dimensions[row].outline_level = 1
            row += 1

        if days_worked > 0:
            extra_h  = total_mins // 60; rem_m = total_mins % 60
            total_hh = total_hours + extra_h
            avg_h    = round(total_hh / days_worked, 1)
            row = _summary_bar(ws, row,
                               f"  {days_worked} days worked   ·   "
                               f"{total_hh}h {rem_m:02d}m total   ·   "
                               f"avg {avg_h}h / day")

        # ── Drawing assignments ───────────────────────────────────────────
        row = _section_label(ws, row, "Drawing Assignments")
        row = _col_headers(ws, row, {
            2: "Project", 3: "Drawing", 4: "Status",
            5: "Done", 6: "Total", 7: "Progress", 8: "Payment",
        })

        projects_for_user = {}
        for project_id, project in projects_all.items():
            au = project.get("assigned_users", {})
            if not isinstance(au, dict) or user_id not in au:
                continue
            proj_drawings = drawings_all.get(project_id, {})
            if not isinstance(proj_drawings, dict):
                continue
            drawings_list = []
            for drawing_id, drawing in proj_drawings.items():
                d_status  = drawing.get("status", "not_started")
                sub_steps = drawing.get("sub_steps", {})
                total_s   = len(sub_steps)
                done_s    = sum(1 for s in sub_steps.values() if s.get("completed", False))
                progress  = f"{done_s}/{total_s}  ({round(done_s/total_s*100)}%)" if total_s else "—"
                paid      = drawing.get("payment_received", False)
                pay_txt   = ("Paid ✓" if paid else "Unpaid ✗") \
                            if d_status in ("client_approved", "admin_approved") else "—"
                drawings_list.append({
                    "drawing":  drawing.get("name", ""),
                    "status":   d_status,
                    "done":     done_s,
                    "total":    total_s,
                    "progress": progress,
                    "pay_txt":  pay_txt,
                    "paid":     paid if d_status in ("client_approved", "admin_approved") else None,
                })
            drawings_list.sort(key=lambda x: STATUS_ORDER.get(x["status"], 9))
            projects_for_user[project_id] = {
                "name":     project.get("name", ""),
                "client":   project.get("client_name", ""),
                "drawings": drawings_list,
            }

        if not projects_for_user:
            _fill_row(ws, row, SURFACE, ROW_BORDER)
            c = ws.cell(row=row, column=2, value="No drawings assigned")
            c.font = Font(name="Calibri", italic=True, color="9CA3AF", size=9)
            c.fill = SURFACE; c.border = ROW_BORDER
            c.alignment = Alignment(vertical="center")
            _merge(ws, row, 2, 8)
            ws.row_dimensions[row].height = 18
            ws.row_dimensions[row].outline_level = 1
            row += 1
        else:
            sorted_projs = sorted(projects_for_user.values(),
                                  key=lambda p: p["name"].lower())
            d_idx = 0
            for proj in sorted_projs:
                for drw in proj["drawings"]:
                    bg = BG_ALT if d_idx % 2 else SURFACE
                    _fill_row(ws, row, bg, ROW_BORDER)
                    _cell(ws, row, 2, f"{proj['name']}  ·  {proj['client']}",
                          fg="64748B", size=9, bg=bg)
                    _cell(ws, row, 3, drw["drawing"],
                          fg="1F2328", size=10, bold=True, bg=bg)
                    _status_cell(ws, row, 4, drw["status"], bg)
                    _cell(ws, row, 5, drw["done"],     bg=bg, ah="center", fg="374151")
                    _cell(ws, row, 6, drw["total"],    bg=bg, ah="center", fg="374151")
                    _cell(ws, row, 7, drw["progress"], bg=bg, ah="center", fg="374151")
                    _pay_cell(ws, row, 8,
                              drw["pay_txt"],
                              drw["paid"] if drw["paid"] is not None else False,
                              bg)
                    ws.row_dimensions[row].height = 20
                    ws.row_dimensions[row].outline_level = 1
                    row += 1
                    d_idx += 1

        row = _gap(ws, row, 14)
        row = _divider(ws, row)
        row = _gap(ws, row, 6)

    # ── Save ─────────────────────────────────────────────────────────────────
    date_range = ""
    if start_date and end_date:
        date_range = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_range = f"_from_{start_date}"
    elif end_date:
        date_range = f"_until_{end_date}"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"employee_performance{date_range}_{timestamp}.xlsx"
    full_path = os.path.join(get_reports_directory(), filename)
    wb.save(full_path)
    logger.info(f"Generated employee performance report: {full_path}")

    # Auto-open in Excel / default spreadsheet app
    try:
        os.startfile(full_path)
    except Exception:
        pass

    return full_path
